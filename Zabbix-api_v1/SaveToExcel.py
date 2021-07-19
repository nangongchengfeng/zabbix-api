#! /usr/bin/env python
# _*_ coding: utf-8 _*_


import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Side, Border, PatternFill


def WriteExcel(FilaPath, ZabbixData):
    WorkBook = Workbook()
    Sheet = WorkBook.active
    Sheet.title = '服务器资源使用情况'
    #  除去 :  '根目录总量/G','根目录使用量/G',
    TableTitle = ['IP','主机名','运行时长/天','CPU/核','内存/GB','根目录使用率/%','CPU平均负载/1min','CPU平均负载/5min','CPU平均负载/15min','CPU空闲时间','CPU使用率/%','内存使用率/%','可用内存/G','磁盘使用率/%','低负载（是/否）']
    TitleColumn = {} #存放每个title值所对应的列{'IP': 'A', '主机名': 'B', '运行时长': 'C', 'CPU/核': 'D', '内存/GB': 'E', '根目录总量': 'F',...}
    AllHostItemValues = [] #存放所有主机的监控项值 列表信息。

    # 维护表头，写入表头数据
    for row in range(len(TableTitle)):
        Col = row + 1
        Column = Sheet.cell(row=1, column=Col)    #获取单元格的位置
        Column.value = TableTitle[row]  #写入数据
        TitleCol = Column.coordinate.strip('1') #获取Title所在的列
        TitleColumn[TableTitle[row]] = TitleCol #加入到TitleColumn

    # 整理Zabbix 监控数据逐行写入到表格中
    #print(ZabbixData)
    for host in ZabbixData.values():
        # 1.首先要对分区目录使用率进行一个整合，将除/目录外的分区目录使用率整合为一个值
        DiskItems = ''   #定义一个空值，用于存放除根目录空间使用率外所有的分区目录使用率
        DelItems = []    #定义一个空列表，用于存放除根目录空间使用率外所有的分区目录使用率的键值
        for item in host:
            DiskItem = re.findall(r'^/[a-z0-9]{1,50}: Space utilization', item)
            if len(DiskItem) == 1:
                DiskItem = DiskItem[0]  #获取监控项的名字 /boot: Space utilization
                NewDiskItem = DiskItem.strip('Space utilization')  # 将名字格式化，/boot: Space utilization 格式化为：/boot:
                DiskItemValue = str(round(float(host[item]), 2)) + '%'  # 取出对应监控项的值，并格式化保留两位小数
                # 将所有分区目录使用率组合为一个整的磁盘使用率
                if DiskItems == '':
                    DiskItemData = str(NewDiskItem) + ' ' + str(DiskItemValue)
                else:
                    DiskItemData = '\n' + str(NewDiskItem) + ' ' + str(DiskItemValue)
                DiskItems += DiskItemData
                # 将处理完的磁盘使用率加入到DelItems列表中，供后续删除使用
                DelItems.append(DiskItem)
        #print(host)
        # 2.将已经整合过的分区目录使用率监控项在原来的主机监控项中删除
        for delitem in DelItems:
            host.pop(delitem)

        # 3.将整合好的分区目录使用率，重新加入到主机监控项的字典中
        host['Disk utilization'] = DiskItems
        #print(host)
        # 4.将每台主机监控项的值取出来组成一个列表
        # 最终得到一条一条这样的数据：
        #'IP','主机名','运行时长/天','CPU/核','内存/GB','根目录使用率/%','CPU平均负载/1min','CPU平均负载/5min','CPU平均负载/15min','CPU空闲时间','CPU使用率/%','内存使用率/%','可用内存/G','磁盘使用率/%'
        # ['172.24.125.12', 'tbds-172-24-125-12', '245.87d', '16', 64, '50G', 7.43, '14.87%', '0.1', '0.18', '0.32', 97.79, '2.21%', '35.52%', 40.45, '/boot: 14.23%\n/data: 6.24%\n/home: 0.03%']
        #print(host['System uptime'])
        if 'System uptime' in host:
            HostItemValues = [] #定义一个空列表，用于存放主机的监控项的值
            HostItemValues.append(host['ip'])
            HostItemValues.append(host['name'])
            try:
                HostItemValues.append(str(round(int(host['System uptime']) / 24 / 60 / 60, 2)) + 'd')
                # 首先将运行时长换算为天数，然后再加入到列表中
            except  IndexError as e:
                print("IndexError Details : " + str(e))
                pass

            HostItemValues.append(host['Number of CPUs'])
            TotalMemory = int(int(host['Total memory']) / 1024 / 1024 / 1024)
            if TotalMemory == 7:
                TotalMemory = 8
            elif TotalMemory == 15:
                TotalMemory = 16
            elif TotalMemory == 31:
                TotalMemory = 32
            elif TotalMemory == 62:
                TotalMemory = 64
            elif TotalMemory == 251:
                TotalMemory = 256
            elif TotalMemory == 503:
                TotalMemory = 512
            HostItemValues.append(TotalMemory)  # 内存总大小
            #HostItemValues.append(str(round(int(host['/: Total space']) / 1024 / 1024 / 1024)) + 'G')  # 根目录总共大小
            #HostItemValues.append(str(round(int(host['/: Used space']) / 1024 / 1024 / 1024, 2)) + 'G')  # 根目录使用量
            HostItemValues.append(str(round(float(host['根目录使用率监控']), 2)) + '%')  # 根目录使用率
            HostItemValues.append(host['Load average (1m avg)'])
            HostItemValues.append(host['Load average (5m avg)'])
            HostItemValues.append(host['Load average (15m avg)'])
            HostItemValues.append(round(float(host['idle time']), 2))  # CPU空闲时间
            HostItemValues.append(str(round(float(host['CPU utilization']), 2)) + '%')  # CPU使用率
            HostItemValues.append(str(round(float(host['Memory utilization']), 2)) + '%')  # 内存使用率
            HostItemValues.append(str(round(int(host['Available memory']) / 1024 / 1024 / 1024, 2)) + 'G')  # 可用内存
            HostItemValues.append(host['服务器硬盘总使用率'])  # 磁盘使用率
            if host['Number of CPUs'] == "0":
                HostItemValues.append("已停用")
            #print(type(float(host['Load average (15m avg)'])),type(float(round(float(host['CPU utilization']), 2))))
            if  float(host['Load average (15m avg)']) >= 10 or  float(round(float(host['CPU utilization']), 2)) >= 20 :
                #print("负载: 是" + host['Load average (15m avg)'])
                #print("cpu: 是" + str(round(float(host['CPU utilization']), 2)))
                HostItemValues.append("否")
            else:
                #print("负载: 否" + host['Load average (15m avg)'])
               # print("cpu: 否" + str(round(float(host['CPU utilization']), 2)))
                HostItemValues.append("是")
            # 将每一台主机的所有监控项信息添加到AllHostItems列表中
            AllHostItemValues.append(HostItemValues)
        #print(AllHostItemValues)
    # 将所有信息写入到表格中
    for HostValue in range(len(AllHostItemValues)):
        Sheet.append(AllHostItemValues[HostValue])
        #print(HostValue)
    ############ 设置单元格样式 ############
    # 字体样式
    TitleFont = Font(name="宋体", size=12, bold=True, italic=False, color="000000")
    TableFont = Font(name="宋体", size=11, bold=False, italic=False, color="000000")
    # 对齐样式
    alignment = Alignment(horizontal="center", vertical="center", text_rotation=0, wrap_text=True)
    # 边框样式
    side1 = Side(style='thin', color='000000')
    border = Border(left=side1, right=side1, top=side1, bottom=side1)
    # 填充样式
    pattern_fill = PatternFill(fill_type='solid', fgColor='99ccff')
    # 设置列宽
    column_width = {'A': 15, 'B': 30, 'C': 14, 'D': 10, 'E': 10, 'F': 16, 'G': 18, 'H': 18, 'I': 22, 'J': 22, 'K': 23,
                    'L': 15, 'M': 16, 'N': 16, 'O': 14, 'P': 16}
    for i in column_width:
        Sheet.column_dimensions[i].width = column_width[i]
    # 设置首行的高度
    Sheet.row_dimensions[1].height = 38
    # 冻结窗口
    Sheet.freeze_panes = 'A2'
    # 添加筛选器
    Sheet.auto_filter.ref = Sheet.dimensions

    # 设置单元格字体及样式
    for row in Sheet.rows:
        for cell in row:
            if cell.coordinate.endswith('1') and len(cell.coordinate) == 2:
                cell.alignment = alignment  #设置对齐样式
                cell.font = TitleFont   #设置字体
                cell.border = border    #设置边框样式
                cell.fill = pattern_fill    #设置填充样式
            else:
                cell.font = TableFont
                cell.alignment = alignment
                cell.border = border
    WorkBook.save(filename=FilaPath)